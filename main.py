"""
To do:
    
"""
import openpyxl, pyautogui, time

#Files
wb = openpyxl.load_workbook('readTest.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
sampleImage = 'testImageSS.png'

#Custom
tabWidth = 34
cursorMoveTime = 0.5
typeInterval = 0.1

#DataStructureInit
jnl = {}
initTime = time.time()

def readJnl(startOfJnl, endOfJnl):
    #Prints limits of jnl if uncommented
    #print(str(startOfJnl) + " " + str(endOfJnl))
    if endOfJnl - startOfJnl == 0:
        return
    else:
        for i in range(startOfJnl, endOfJnl):
            if sheet.cell(row = i, column = 6).value == sheet.cell(row = 1, column = 1).value:
                jnl[sheet.cell(row = i, column = 4).value] = 0
            else:
                jnl[sheet.cell(row = i, column = 4).value] = sheet.cell(row = i, column = 6).value
    

def writeJnl(jnlNo):
    cellCoords = pyautogui.locateOnScreen(sampleImage)
    #Move cursor to starting input cell
    pyautogui.moveTo((cellCoords[2]),cellCoords[1]+((cellCoords[3]/5)*2.75), cursorMoveTime)
    pyautogui.click(clicks = 1, button = 'left')
    pyautogui.typewrite(str(jnlNo), interval = typeInterval)
    pyautogui.typewrite(['right', 'right', 'enter', 'enter'], interval = typeInterval)
    for key in jnl:
        pyautogui.typewrite(str(jnl[key]), interval = typeInterval)
        pyautogui.typewrite(['enter'], interval = typeInterval)

def main():
    #Setting defs: count for checking if full jnl found, others to avoid called b4 defined
    count = 0
    startOfJnl = 0
    endOfJnl = 0
    jnlNo = 0
    #Loops through each row to find jnls (uses +1 since row references start at 1)
    for rowNum in range(1, sheet.max_row+1):
        #Check if jnlno col is filled (i.e start or end of jnl)
        if sheet.cell(row = rowNum, column = 2).value != sheet.cell(row = 1, column = 1).value:
            count += 1
            #Checks to see if it is the end or start of jnl
            if count%2==0:
                endOfJnl = rowNum
                readJnl(startOfJnl, endOfJnl)
                writeJnl(jnlNo)
            else:
                startOfJnl = rowNum
                jnlNo = sheet.cell(row = rowNum, column = 2).value
    elapsedTime = time.time()-initTime
    print(jnl)
    print('Time to complete = ' + str(elapsedTime))
    
    
""" COUNTDOWN
for i in range(3):
    time.sleep(1)
    print(3-i)
"""

main()

"""Use pyautogui.locateOnScreen('screenshot of cell to input the data into
then use the pyautogui.moveRel(xOffset, yOffset, duration=numofsecs)
"""

